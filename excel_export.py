from __future__ import annotations

import io
import json
import logging
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Case ID column olib tashlandi
COLUMNS = [
    ("№",            4),
    ("Ish raqami",  22),
    ("Majlis sanasi", 14),
    ("Vaqt",         8),
    ("Sud",          32),
    ("Instansiya",  20),
    ("Kategoriya",  35),
    ("Da'vogar",    45),
    ("Javobgar",    45),
    ("Sudya",       30),
]

HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
EVEN_FILL    = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
ODD_FILL     = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
NEW_FILL     = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
UPDATED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@lru_cache(maxsize=1)
def _build_court_map() -> dict[str, str]:
    """regions.json dan {value -> name} map, bir marta yuklanadi."""
    path = Path(__file__).parent / "regions.json"
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        mapping: dict[str, str] = {}
        for entry in data:
            for region in entry.get("type_court", []):
                for court in region.get("tumanlar", []):
                    v = court.get("value")
                    n = court.get("name")
                    if v and n:
                        mapping[v] = n
        return mapping
    except Exception as exc:
        logger.warning(f"regions.json yuklanmadi: {exc}")
        return {}


def _court_name(globalid: str) -> str:
    """'tashxsud' -> 'Тошкент шаҳар суди'. Topilmasa globalid qaytadi."""
    return _build_court_map().get(globalid, globalid)


def _cell_align(wrap: bool = False) -> Alignment:
    return Alignment(vertical="center", wrap_text=wrap)


def build_excel(
    records: list[dict[str, Any]],
    new_ids: set[str] | None = None,
    updated_ids: set[str] | None = None,
    sheet_title: str = "Sud majlislari",
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.freeze_panes = "A2"

    # Header
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, record in enumerate(records, start=2):
        case_id = record.get("case_id", "")

        if new_ids and case_id in new_ids:
            row_fill = NEW_FILL
        elif updated_ids and case_id in updated_ids:
            row_fill = UPDATED_FILL
        elif row_idx % 2 == 0:
            row_fill = EVEN_FILL
        else:
            row_fill = ODD_FILL

        values = [
            row_idx - 1,
            record.get("casenumber", ""),
            record.get("hearing_date", ""),
            record.get("hearing_time", ""),
            _court_name(record.get("globalid", "")),   # value -> to'liq nom
            record.get("instance", ""),
            record.get("category", ""),
            record.get("claiment", ""),
            record.get("defendant", ""),
            record.get("responsible", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = _cell_align(wrap=(col_idx in (5, 7, 8, 9)))

        ws.row_dimensions[row_idx].height = 18

    # Legend
    if new_ids or updated_ids:
        legend_row = len(records) + 3
        if new_ids:
            c = ws.cell(row=legend_row, column=1, value="🟢 Yangi yozuvlar")
            c.fill = NEW_FILL
            c.font = Font(bold=True, size=10)
        if updated_ids:
            c = ws.cell(row=legend_row + 1, column=1, value="🟡 O'zgargan yozuvlar")
            c.fill = UPDATED_FILL
            c.font = Font(bold=True, size=10)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_new_records_excel(records: list[dict[str, Any]]) -> io.BytesIO:
    new_ids = {r["case_id"] for r in records if r.get("_status") == "new"}
    updated_ids = {r["case_id"] for r in records if r.get("_status") == "updated"}
    return build_excel(records, new_ids=new_ids, updated_ids=updated_ids, sheet_title="Yangi majlislar")


def filename_now(prefix: str = "sud") -> str:
    return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"